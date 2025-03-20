import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import os
import xlrd
import Levenshtein


class DataDesensitizationApp:
    def __init__(self, master):
        self.master = master
        master.title("Excel数据脱敏工具")
        master.geometry("600x410")

        # 创建界面组件
        self.create_widgets()

    def create_widgets(self):
        # 文件选择
        ttk.Label(self.master, text="选择Excel文件:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.file_entry = ttk.Entry(self.master, width=40)
        self.file_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(self.master, text="浏览", command=self.browse_file).grid(row=0, column=2, padx=5, pady=5)

        # 选择脱敏模式（表头 or 正则）
        self.mode_var = tk.StringVar(value="header")  # 默认按表头脱敏
        ttk.Label(self.master, text="脱敏模式:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        ttk.Radiobutton(self.master, text="已知表头脱敏", variable=self.mode_var, value="header").grid(row=1, column=1,
                                                                                                       sticky="w")
        ttk.Radiobutton(self.master, text="正则匹配脱敏", variable=self.mode_var, value="regex").grid(row=1, column=2,
                                                                                                      sticky="w")

        # 敏感词
        ttk.Label(self.master, text="敏感词(逗号分隔):").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.keywords_entry = ttk.Entry(self.master, width=40)
        self.keywords_entry.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        # 选择脱敏规则
        self.rule_var = tk.StringVar(value="all")
        ttk.Label(self.master, text="脱敏规则:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        ttk.Radiobutton(self.master, text="全部替换(*)", variable=self.rule_var, value="all",
                        command=self.toggle_custom_entry).grid(row=3, column=1, sticky="w")
        ttk.Radiobutton(self.master, text="部分替换(*)", variable=self.rule_var, value="partial",
                        command=self.toggle_custom_entry).grid(row=3, column=2, sticky="w")
        ttk.Radiobutton(self.master, text="全部替换(自定义字符)", variable=self.rule_var, value="custom",
                        command=self.toggle_custom_entry).grid(row=4, column=1, sticky="w")
        ttk.Radiobutton(self.master, text="部分替换(自定义字符)", variable=self.rule_var, value="partial_custom",
                        command=self.toggle_custom_entry).grid(row=4, column=2, sticky="w")

        # 自定义字符输入（默认禁用）
        ttk.Label(self.master, text="自定义替换字符:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.custom_char_entry = ttk.Entry(self.master, width=10, state="disabled")
        self.custom_char_entry.grid(row=5, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        # 保存路径
        ttk.Label(self.master, text="保存路径:").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        self.save_entry = ttk.Entry(self.master, width=40)
        self.save_entry.grid(row=6, column=1, padx=5, pady=5)
        ttk.Button(self.master, text="浏览", command=self.browse_save_dir).grid(row=6, column=2, padx=5, pady=5)

        # 开始处理按钮
        ttk.Button(self.master, text="开始脱敏", command=self.process_file).grid(row=7, column=1, columnspan=2, pady=10)
        ttk.Label(self.master, text="选择多个文件用英文分号分隔").grid(row=8, column=1, padx=5, sticky="w")
        ttk.Label(self.master, text="选择已知表头脱敏需要填写敏感词(表头)").grid(row=9, column=1, padx=5, sticky="w")
        ttk.Label(self.master, text="正则匹配脱敏不需要填写敏感词").grid(row=10, column=1, padx=5, sticky="w")
        ttk.Label(self.master, text="正则表达式目前已实装：").grid(row=11, column=1, padx=5, sticky="w")
        ttk.Label(self.master, text="手机号、身份证号、邮箱、IP地址、行政地址、URL网址").grid(row=12, column=1, padx=5, sticky="w")

    def toggle_custom_entry(self):
        """切换自定义字符输入框的启用/禁用状态"""
        if self.rule_var.get() in ["custom", "partial_custom"]:
            self.custom_char_entry.config(state="normal")  # 允许输入
        else:
            self.custom_char_entry.config(state="disabled")  # 禁用输入
            self.custom_char_entry.delete(0, tk.END)  # 清空内容

    @staticmethod
    def levenshtein_similarity(str1, str2):
        distance = Levenshtein.distance(str1, str2)
        max_length = max(len(str1), len(str2))
        return 1 - (distance / max_length)

    @staticmethod
    def jaccard_similarity(str1, str2):
        set1, set2 = set(str1), set(str2)
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        return intersection / union if union != 0 else 0

    @staticmethod
    def compute_similarity(header, keyword):
        lev_sim = DataDesensitizationApp.levenshtein_similarity(header, keyword)
        jacc_sim = DataDesensitizationApp.jaccard_similarity(header, keyword)
        return (lev_sim + jacc_sim) / 2

    def process_regex(self, sheet, rule, custom_char):
        patterns = {  # 正则匹配
            "phone": r"(?<!\d)1[3-9]\d{1}-?\d{4}-?\d{4}(?!\d)",  # 手机号
            "id_card": r"\b\d{18}|\d{17}X\b",  # 身份证号
            "email": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b",  # 邮箱
            "ip": r"\b(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\b",
            # ip地址
            "address": r"(?=.*?(?:省|市|区|县|镇|乡|村))"  # 至少包含一个行政单位
                       r"([\u4e00-\u9fa5]{2,9}(?:省|自治区|市))?"  # 省级
                       r"([\u4e00-\u9fa5]{2,9}(?:市|自治州|区|县|旗))?"  # 市级或区县级
                       r"((?<!(?:隔壁|对面|附近))[\u4e00-\u9fa5]{2,9}(?:街道|镇|乡|村))?"  # 乡镇级（排除干扰词）
                       r"([\u4e00-\u9fa5]{1,8}?(?:号|路|街|巷|大厦|小区|学|院|室|单元))",  # 行政地名
            "url": r"(http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*,]|(?:%[0-9a-fA-F][0-9a-fA-F]))+)|"
                   r"([a-zA-Z]+.\w+\.+[a-zA-Z0-9\/_]+)"  # URL网址
        }

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.value = str(cell.value)
                    if cell.value and isinstance(cell.value, str):
                        for pattern in patterns.values():
                            cell.value = re.sub(pattern, lambda x: self.desensitize(x.group(), rule, custom_char),
                                                cell.value)

    def desensitize(self, text, rule, custom_char="*"):
        """ 对文本进行脱敏处理 """
        if rule == "all":  # 全部替换为 *
            return "*" * len(text)

        elif rule == "partial":  # 部分替换（保留前2后2）
            if len(text) > 4:
                return text[:1] + "*" * (len(text) - 4) + text[-1:]
            else:
                return "*" * len(text)

        elif rule == "custom":  # 全部替换（自定义字符）
            return custom_char * len(text)

        elif rule == "partial_custom":  # 部分替换（自定义字符）
            if len(text) > 4:
                return text[:1] + custom_char * (len(text) - 4) + text[-1:]
            else:
                return custom_char * len(text)

        return text  # 其他情况不变

    def browse_file(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if file_paths:
            self.files = list(file_paths)  # 存储多个文件路径
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, "; ".join(file_paths))  # 在UI中显示

    def browse_save_dir(self):
        dir_path = filedialog.askdirectory()
        if dir_path:
            self.save_entry.delete(0, tk.END)
            self.save_entry.insert(0, dir_path)

    def convert_xls_to_xlsx(self, file_path):
        """将 .xls 文件转换为 .xlsx 并返回 openpyxl Workbook"""
        book = xlrd.open_workbook(file_path)
        wb = Workbook()

        for sheet_index in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_index)
            ws = wb.create_sheet(title=sheet.name if sheet_index > 0 else "Sheet1")

            for row_idx in range(sheet.nrows):
                row_values = sheet.row_values(row_idx)
                ws.append(row_values)

        return wb

    def process_file(self):
        try:
            save_dir = self.save_entry.get()
            keywords = [k.strip() for k in self.keywords_entry.get().split(",")]
            rule = self.rule_var.get()
            custom_char = self.custom_char_entry.get()
            mode = self.mode_var.get()  # 选择的脱敏模式（表头匹配 or 正则匹配）

            if not self.files or not save_dir:
                messagebox.showerror("错误", "请填写所有必填项")
                return

            for file_path in self.files:
                file_extension = file_path.split(".")[-1].lower()

                if file_extension == "xlsx":
                    wb = load_workbook(filename=file_path)
                elif file_extension == "xls":
                    wb = self.convert_xls_to_xlsx(file_path)
                else:
                    messagebox.showerror("错误", f"不支持的文件格式: {file_path}")
                    continue

                for sheet in wb.worksheets:
                    if mode == "header":
                        if self.is_column_based(sheet, keywords):
                            self.process_columns(sheet, keywords, rule)
                        else:
                            self.process_rows(sheet, keywords, rule)
                    elif mode == "regex":
                        self.process_regex(sheet, rule, custom_char)

                file_name = os.path.basename(file_path)
                if file_name.endswith(".xls"):
                    file_name = file_name.replace(".xls", ".xlsx")
                output_path = os.path.join(save_dir, f"desensitized_{file_name}")

                wb.save(output_path)
                messagebox.showinfo("成功", f"文件已保存到:\n{output_path}")

        except Exception as e:
            messagebox.showerror("错误", f"处理过程中发生错误:\n{str(e)}")

    def is_column_based(self, sheet, keywords, threshold=0.7):
        """判断是否按列处理，使用相似度匹配"""
        first_row = [str(cell.value) for cell in sheet[1] if cell.value]

        match_count = 0
        for header in first_row:
            for keyword in keywords:
                if self.compute_similarity(header, keyword) > threshold:
                    match_count += 1
                    break  # 避免重复计算

        return match_count > len(keywords) // 2  # 超过一半匹配才按列处理

    def process_columns(self, sheet, keywords, rule):
        for col in sheet.iter_cols(min_row=1, max_row=1):
            header = col[0].value
            if header and header in keywords:
                for row in sheet.iter_rows(min_row=2, min_col=col[0].column, max_col=col[0].column):
                    cell = row[0]
                    if cell.value is not None:  # 避免填充 None
                        cell.value = self.desensitize(str(cell.value), rule)

    def process_rows(self, sheet, keywords, rule):
        for row in sheet.iter_rows():
            if any(cell.value is not None and str(cell.value) in keywords for cell in row):
                for cell in row:
                    if cell.value is not None:  # 避免填充 None
                        cell.value = self.desensitize(str(cell.value), rule)


if __name__ == "__main__":
    root = tk.Tk()
    app = DataDesensitizationApp(root)
    root.mainloop()
